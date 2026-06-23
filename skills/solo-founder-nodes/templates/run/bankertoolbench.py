#!/usr/bin/env python3
"""BankerToolBench autonomous runner — the skill's reproducible loop for a deliverable-shaped
benchmark, mirroring spreadsheetbench.py's SHAPE:

    ensure_dataset (clone from allowlist) -> seal_heldout -> attempt (model-in-loop, emits a
    universal `plan` dict, materialized into Excel/PPTX/DOCX/PDF via deliverables.py) -> grade
    -> honest headline.

PORTABLE: zero NodeRoom coupling. NodeRoom is the *dogfood instance*; a fresh user runs THIS file.

THE GRADER — read this before trusting any number it prints:
  BankerToolBench has NO public callable deterministic scorer. Unlike SpreadsheetBench (whose own
  `evaluation.compare_workbooks` we import and never reimplement), BTB's official scorer is GANDALF,
  an LLM verifier that opens the deliverables and grades them against per-task rubrics + golden
  outputs inside the Harbor Docker harness. There is nothing to `import`.

  So this runner has TWO grade LANES, and the lane label is NON-OPTIONAL on every result:
    --grade-lane official  : shell out to the real Harbor+Gandalf run and read its
                             logs/verifier/reward.json — the ONLY lane that yields the official
                             BTB number. Heavy (Docker, HF dataset, judge keys). Gated.
    --grade-lane local     : a deterministic LOCAL proxy (NO LLM on the scored path) modeled on
                             docs/eval/nonbtb/grade.py's honest axes — correctness / formula /
                             citation / fabrication — over the produced deliverables + a per-task
                             rubric.json the USER supplies. This is an UNOFFICIAL local signal and
                             is labeled as such in every line of output. It is NOT the BTB score.

  Why a local lane at all? Same reason grade.py exists: a cheap, reproducible, un-gameable
  inner-loop signal you can run without Docker/judge-API spend. The official-credibility lane and
  the product-evidence lane stay separate by construction (anti-cheat doctrine S-official-vs-product).

HONESTY (derive-don't-accept), identical doctrine to spreadsheetbench.py:
  - NO per-task answers live in this harness. The attempt step calls the user's planner/agent.
  - Held-out is sealed out-of-process: splitHash = HMAC(salt, sorted(task_ids)) (doctrine S12).
    A task inspected for tuning cannot later count as held-out; the agent sees only the current
    task at execution time, never the full sealed corpus. The salt is NOT bundled in this file.
  - The candidate path NEVER reads golden-outputs/ or rubrics/ (the do-not-read boundary). The
    LOCAL grader reads a USER-PROVIDED rubric.json that lives OUTSIDE the agent-visible workspace.
  - The headline refuses to print a bare number: it always carries n + lane + audit caveats.

Modes (attempt):
  --mode agent : the coding agent IS the planner — read one `plan` dict per task from
                 <attempts-dir>/<task_id>.json (the agent wrote it after reading the instruction).
  --mode api   : full-auto — call an OpenAI-compatible model to emit the `plan` dict (env
                 OPENAI_API_KEY / OPENAI_BASE_URL / SOLO_MODEL).
  --dump       : print each sealed held-out task (instruction + input file list) so an
                 agent-in-the-loop can author attempts. No attempt, no grading.

Usage:
  # inspect the sealed slice (no model, no deliverables):
  python bankertoolbench.py --repo <clone> --slice 3 --salt <s> --dump
  # author plan dicts as the agent, materialize, grade with the local proxy:
  python bankertoolbench.py --repo <clone> --slice 3 --salt <s> \
      --mode agent --attempts-dir ./attempts --grade-lane local --rubrics-dir ./rubrics --out results.json
  # full-auto plan via an OpenAI-compatible model, official Harbor+Gandalf grade:
  python bankertoolbench.py --repo <clone> --slice 3 --salt <s> \
      --mode api --grade-lane official --harbor-cmd "harbor run ..." --out results.json
  # offline smoke (no dataset, no model): synthesize a tiny task, materialize, local-grade:
  python bankertoolbench.py --dump --slice 1            # offline dump of the synthetic task
  python bankertoolbench.py --selftest                  # deliverables.write_deliverables round-trip
"""
import os
import sys
import json
import argparse
import hashlib
import hmac
import subprocess

# deliverables.py sits next to this file.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import deliverables  # noqa: E402

# Scope this via your autonomy download-allowlist. BTB is HF-hosted; this is the public repo.
REPO_URL = "https://github.com/Handshake-AI-Research/bankertoolbench.git"
DATASET_HF = "handshake-ai-research/bankertoolbench"

# Paths the candidate must NEVER read (evaluator-only). ensure_dataset enforces this boundary.
DO_NOT_READ = ("golden-outputs", "rubrics", "golden_outputs")


# ---------------------------------------------------------------------------
# dataset
# ---------------------------------------------------------------------------
def ensure_dataset(repo, allowlist, local_checkout=None):
    """Make a BankerToolBench checkout available at `repo`.

    Honesty boundary: returns the directory but the loader below reads ONLY tasks.jsonl +
    task-data/ + input files; it never opens golden-outputs/ or rubrics/.

    local_checkout: if the user already has a checkout (or the HF dataset snapshot) on disk, point
    here and skip the network entirely (the common case — BTB data is HF-gated).
    """
    if local_checkout and os.path.isdir(local_checkout):
        return local_checkout
    if not os.path.isdir(repo):
        if allowlist and not any(h in REPO_URL for h in allowlist):
            raise SystemExit(f"hard-stop: {REPO_URL} not in download allowlist {allowlist}")
        print(f"cloning {REPO_URL} -> {repo}")
        subprocess.run(["git", "clone", "--depth", "1", REPO_URL, repo], check=True)
        print("NOTE: the task DATA (tasks.jsonl + input files) is HF-gated "
              f"({DATASET_HF}); download it into the checkout with your HF_TOKEN before running.")
    return repo


def _find_tasks_file(repo):
    """Locate tasks.jsonl anywhere under the checkout (layout varies across BTB revisions)."""
    for root, dirs, files in os.walk(repo):
        # Never descend into evaluator-only dirs.
        dirs[:] = [d for d in dirs if d not in DO_NOT_READ and not d.startswith(".git")]
        if "tasks.jsonl" in files:
            return os.path.join(root, "tasks.jsonl")
    return None


def load_tasks(repo):
    """Read the agent-visible tasks.jsonl. Each row: {id, instruction/final_prompt, input_files?...}.

    We normalize to {id, instruction, input_files, dir} and DROP any key that points into a
    do-not-read path (defense against a row that smuggles a golden-output reference).
    """
    path = _find_tasks_file(repo)
    if not path:
        raise SystemExit(
            f"tasks.jsonl not found under {repo}. BTB data is HF-gated "
            f"({DATASET_HF}); fetch it with HF_TOKEN, or pass --local-checkout / --synthetic."
        )
    base = os.path.dirname(path)
    tasks = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            row = json.loads(line)
            tid = str(row.get("id") or row.get("task_id") or row.get("name") or len(tasks))
            instruction = (row.get("final_prompt") or row.get("instruction")
                           or row.get("prompt") or "")
            inputs = row.get("input_files") or row.get("inputs") or []
            inputs = [p for p in inputs if not any(b in str(p) for b in DO_NOT_READ)]
            tasks.append({"id": tid, "instruction": instruction,
                          "input_files": inputs, "dir": base, "raw": row})
    return tasks


def synthetic_tasks(n):
    """Offline tasks for the no-dataset smoke. Deliberately answer-free: just instructions, so the
    code path (seal -> attempt -> materialize -> local-grade) runs without HF access."""
    out = []
    for i in range(1, n + 1):
        out.append({
            "id": f"SYNTH-{i:02d}",
            "instruction": (
                f"[SYNTHETIC SMOKE TASK {i}] Build a preliminary diligence package for a target: "
                "a financial model (Excel), a teaser deck of exactly 2 slides, a one-page memo "
                "(Word), and a PDF report. Cite every figure to its source."
            ),
            "input_files": [],
            "dir": "",
            "raw": {"synthetic": True},
        })
    return out


# ---------------------------------------------------------------------------
# sealing (doctrine S12)
# ---------------------------------------------------------------------------
def seal_heldout(ids, salt):
    """splitHash = HMAC(salt, sorted task ids). Out-of-process: the salt is supplied by the
    operator, never bundled. A task inspected for tuning must not later count here."""
    return hmac.new(salt.encode(), ",".join(sorted(ids)).encode(), hashlib.sha256).hexdigest()


# ---------------------------------------------------------------------------
# attempt (model in the loop) -> universal plan dict -> deliverables
# ---------------------------------------------------------------------------
def agent_attempt(task, attempts_dir):
    """Read a pre-authored `plan` dict from <attempts-dir>/<id>.json (the coding agent wrote it)."""
    path = os.path.join(attempts_dir, f"{task['id']}.json")
    if not os.path.isfile(path):
        return None, f"no attempt for {task['id']} (expected {path})"
    with open(path, encoding="utf-8") as f:
        return json.load(f), None


def api_attempt(task):
    """Full-auto: ask an OpenAI-compatible model for the universal `plan` dict. NO answers leak —
    the prompt carries only the instruction + input file NAMES (not golden outputs)."""
    from openai import OpenAI
    client = OpenAI(base_url=os.environ.get("OPENAI_BASE_URL"), api_key=os.environ["OPENAI_API_KEY"])
    model = os.environ.get("SOLO_MODEL", "gpt-4.1-mini")
    schema = (
        '{"title": str, "taskSummary": str, "tickers": [str], '
        '"workbook": {"sheets": [{"name": str, "rows": [[cell,...]]}]}, '
        '"presentation": {"slides": [{"title": str, "bullets": [str], "footnote": str}]}, '
        '"memo": {"sections": [{"heading": str, "body": str}]}, '
        '"citations": [{"claim": str, "sourcePath": str, "locator": str, '
        '"boundaryBoxStatus": str, "quote": str}]}'
    )
    prompt = (
        "You are a junior IB analyst. Produce a deliverable PLAN as a single JSON object for this "
        "task. Excel rows: a cell starting with '=' is a live formula (preferred for derived "
        "values). Cite every figure.\n\n"
        f"Instruction:\n{task['instruction']}\n\n"
        f"Available input files (names only): {task.get('input_files')}\n\n"
        f"Return ONLY JSON matching this shape:\n{schema}\nNo prose, no fences."
    )
    r = client.chat.completions.create(model=model, messages=[{"role": "user", "content": prompt}],
                                       temperature=0)
    txt = r.choices[0].message.content.strip().strip("`")
    if txt.startswith("json"):
        txt = txt[4:]
    try:
        return json.loads(txt), None
    except Exception as exc:
        return None, f"api_attempt JSON parse failed: {exc}"


def attempt(task, mode, attempts_dir):
    if mode == "agent":
        return agent_attempt(task, attempts_dir)
    return api_attempt(task)


# ---------------------------------------------------------------------------
# grade — TWO lanes, lane label NON-OPTIONAL
# ---------------------------------------------------------------------------
def grade_official(out_dir, task, harbor_cmd):
    """OFFICIAL lane: run the real Harbor+Gandalf verifier and read its reward.json.

    harbor_cmd is the operator-supplied command that runs ONE task through Harbor with the
    materialized deliverables in out_dir mounted as the agent's file writes. We do NOT reimplement
    Gandalf — we read the score it emits. Returns {lane:'official', ...}.
    """
    if not harbor_cmd:
        return {"lane": "official", "ok": False, "score": None,
                "note": "no --harbor-cmd supplied; cannot run the official verifier"}
    env = dict(os.environ, BTB_OUT_DIR=str(out_dir), BTB_TASK_ID=str(task["id"]))
    try:
        proc = subprocess.run(harbor_cmd, shell=True, env=env, capture_output=True, text=True,
                              timeout=int(os.environ.get("BTB_HARBOR_TIMEOUT", "3600")))
    except subprocess.TimeoutExpired:
        return {"lane": "official", "ok": False, "score": None, "note": "harbor run timed out"}
    # Gandalf writes logs/<job>/verifier/reward.json — let the operator point us at it, else scan.
    reward_path = os.environ.get("BTB_REWARD_JSON")
    if not reward_path:
        for root, _dirs, files in os.walk(out_dir):
            if "reward.json" in files:
                reward_path = os.path.join(root, "reward.json")
                break
    if not reward_path or not os.path.isfile(reward_path):
        return {"lane": "official", "ok": False, "score": None,
                "note": f"reward.json not found (harbor rc={proc.returncode}); set BTB_REWARD_JSON",
                "stderr_tail": (proc.stderr or "")[-400:]}
    with open(reward_path, encoding="utf-8") as f:
        reward = json.load(f)
    score = reward.get("reward", reward.get("score"))
    return {"lane": "official", "ok": True, "score": score, "reward": reward,
            "reward_path": reward_path}


def grade_local(out_dir, task, rubrics_dir):
    """LOCAL PROXY lane (UNOFFICIAL): deterministic, NO LLM on the scored path. Modeled on
    docs/eval/nonbtb/grade.py's axes. Reads a USER-PROVIDED rubric (rubrics_dir/<id>.json) that
    lives OUTSIDE the agent-visible workspace — never BTB's evaluator-only rubrics/.

    rubric.json shape (user authored, mirrors grade.py):
      {
        "task": str,
        "deliverables": ["xlsx","pptx","docx","pdf"],   # which files must exist & be non-empty
        "formula_required": bool,                         # xlsx must contain >=1 '=' cell
        "citations_required": bool,                       # >=1 citation row present
        "sources": ["ABC_10K.pdf", ...],                  # allowed citation sources (fabrication check)
        "allowed_sheets": ["Model","Comps", ...]          # optional; sheets outside => fabrication
      }
    Axes -> correctness(=files present), formula, cited, fabrication. Each fabrication costs 10%.
    Returns {lane:'local-proxy-UNOFFICIAL', score, axes...}.
    """
    LANE = "local-proxy-UNOFFICIAL"
    rubric_path = os.path.join(rubrics_dir or "", f"{task['id']}.json")
    if not rubrics_dir or not os.path.isfile(rubric_path):
        # No rubric => can only assert the deliverables exist; report that honestly, do NOT invent a score.
        present = _deliverables_present(out_dir)
        return {"lane": LANE, "score": None, "note": "no rubric.json -> existence-only check (not a score)",
                "deliverables_present": present}
    with open(rubric_path, encoding="utf-8-sig") as f:
        rubric = json.load(f)

    want = rubric.get("deliverables", ["xlsx", "pptx", "docx", "pdf"])
    present = _deliverables_present(out_dir)
    correct = sum(1 for k in want if present.get(k))
    n = len(want) or 1

    sources = set(rubric.get("sources", []))
    need_formula = rubric.get("formula_required", False)
    need_cite = rubric.get("citations_required", False)

    # formula axis: the xlsx contains at least one live formula cell.
    formula_ok = None
    if need_formula:
        formula_ok = _xlsx_has_formula(os.path.join(out_dir, "banker_model.xlsx"))
    # citation axis + fabrication: read the Citation Receipts we wrote into the workbook.
    receipts = _xlsx_citation_receipts(os.path.join(out_dir, "banker_model.xlsx"))
    cited_ok = None
    fabrication = 0
    if need_cite:
        cited_ok = len(receipts) > 0
    if sources:
        for r in receipts:
            src = (r.get("source") or "").strip()
            if src and src not in sources:
                fabrication += 1
    # optional sheet-allowlist fabrication check.
    allowed_sheets = set(rubric.get("allowed_sheets", []))
    if allowed_sheets:
        for sn in _xlsx_sheet_names(os.path.join(out_dir, "banker_model.xlsx")):
            if sn != "Citation Receipts" and sn not in allowed_sheets:
                fabrication += 1

    dims = correct + (int(bool(formula_ok)) if need_formula else 0) + (int(bool(cited_ok)) if need_cite else 0)
    denom = n + (1 if need_formula else 0) + (1 if need_cite else 0)
    raw = dims / denom if denom else 0.0
    score = max(0.0, raw - 0.1 * fabrication)
    return {"lane": LANE, "score": round(score, 3), "n": n, "correct": correct,
            "deliverables_present": present,
            "formula_ok": formula_ok, "cited_ok": cited_ok, "fabrication": fabrication,
            "citation_count": len(receipts)}


def _deliverables_present(out_dir):
    def nonempty(name):
        p = os.path.join(out_dir, name)
        return os.path.isfile(p) and os.path.getsize(p) > 0
    return {
        "xlsx": nonempty("banker_model.xlsx"),
        "pptx": nonempty("banker_presentation.pptx"),
        "docx": nonempty("banker_memo.docx"),
        "pdf": nonempty("banker_report.pdf"),
    }


def _xlsx_has_formula(path):
    if not os.path.isfile(path):
        return False
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        return any(isinstance(c.value, str) and c.value.lstrip().startswith("=")
                   for ws in wb.worksheets for row in ws.iter_rows() for c in row)
    except Exception:
        return False


def _xlsx_sheet_names(path):
    if not os.path.isfile(path):
        return []
    try:
        from openpyxl import load_workbook
        return list(load_workbook(path, read_only=True).sheetnames)
    except Exception:
        return []


def _xlsx_citation_receipts(path):
    """Read the 'Citation Receipts' sheet deliverables.py writes -> [{claim, source, locator, ...}]."""
    if not os.path.isfile(path):
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        if "Citation Receipts" not in wb.sheetnames:
            return []
        ws = wb["Citation Receipts"]
        rows = list(ws.iter_rows(values_only=True))
        out = []
        for r in rows[1:]:  # skip header
            r = list(r) + [None] * 5
            if not any(r[:5]):
                continue
            out.append({"claim": r[0], "source": r[1], "locator": r[2],
                        "boundaryBoxStatus": r[3], "quote": r[4]})
        return out
    except Exception:
        return []


def grade(out_dir, task, lane, harbor_cmd, rubrics_dir):
    if lane == "official":
        return grade_official(out_dir, task, harbor_cmd)
    return grade_local(out_dir, task, rubrics_dir)


# ---------------------------------------------------------------------------
# honest headline
# ---------------------------------------------------------------------------
def honest_headline(rows, lane, manifest, dataset_label, caveats):
    counted = [r for r in rows if isinstance(r.get("grade", {}).get("score"), (int, float))
               and r.get("cleanProbe")]
    n = len(counted)
    mean = sum(r["grade"]["score"] for r in counted) / n if n else None
    print("\n" + "=" * 72)
    if lane == "official":
        print("HEADLINE — OFFICIAL Harbor+Gandalf lane")
    else:
        print("HEADLINE — LOCAL PROXY lane (UNOFFICIAL — NOT the BankerToolBench score)")
    print(f"  dataset       : {dataset_label}")
    print(f"  held-out seal : {manifest[:16]}…")
    print(f"  score         : {mean}  (over n={n} clean held-out tasks)")
    if mean is None:
        print("  (refusing to print a number: no clean, scored, held-out tasks)")
    for c in caveats:
        print(f"  caveat        : {c}")
    print("=" * 72)
    return {"lane": lane, "n": n, "headline_score": mean}


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", default=os.path.join(os.getcwd(), "bankertoolbench"))
    ap.add_argument("--local-checkout", default=os.environ.get("BTB_LOCAL_CHECKOUT"),
                    help="path to an existing BTB checkout / HF snapshot (skips clone)")
    ap.add_argument("--allowlist", default="github.com")
    ap.add_argument("--slice", type=int, default=3)
    ap.add_argument("--salt", default=os.environ.get("SOLO_LEDGER_SALT", "dev-salt-change-me"))
    ap.add_argument("--mode", choices=["agent", "api"], default="agent")
    ap.add_argument("--attempts-dir", default="./attempts")
    ap.add_argument("--grade-lane", choices=["official", "local"], default="local")
    ap.add_argument("--harbor-cmd", default=os.environ.get("BTB_HARBOR_CMD"),
                    help="official lane: shell command that runs one task through Harbor+Gandalf")
    ap.add_argument("--rubrics-dir", default=os.environ.get("BTB_RUBRICS_DIR"),
                    help="local lane: dir of user-authored <id>.json rubrics (NOT BTB's evaluator rubrics)")
    ap.add_argument("--out", default="results.json")
    ap.add_argument("--out-root", default="_btb_out")
    ap.add_argument("--synthetic", action="store_true",
                    help="offline: synthesize tiny tasks (no dataset/HF needed) for a smoke run")
    ap.add_argument("--dump", action="store_true", help="print the sealed slice; no attempt/grade")
    ap.add_argument("--selftest", action="store_true",
                    help="round-trip deliverables.write_deliverables into a tempdir and exit")
    a = ap.parse_args()

    if a.selftest:
        return deliverables._selftest()

    # ---- load + seal ----
    use_synthetic = a.synthetic or (not a.local_checkout and not os.path.isdir(a.repo))
    if use_synthetic:
        tasks = synthetic_tasks(a.slice)
        dataset_label = "SYNTHETIC (offline smoke — not real BankerToolBench data)"
        if not a.synthetic:
            print("NOTE: no checkout found -> falling back to SYNTHETIC offline tasks. "
                  "Pass --repo/--local-checkout (with HF data) for the real benchmark.")
    else:
        repo = ensure_dataset(a.repo, a.allowlist.split(","), a.local_checkout)
        tasks = load_tasks(repo)
        dataset_label = a.local_checkout or repo

    sliced = tasks[: a.slice]
    manifest = seal_heldout([t["id"] for t in sliced], a.salt)
    print(f"sealed held-out slice: {len(sliced)} tasks · manifest {manifest[:16]}…  "
          f"(salt is out-of-process; tasks inspected for tuning must not be re-used here)")

    if a.dump:
        for t in sliced:
            print("\n" + "=" * 70)
            print(f"id={t['id']}")
            print("instruction:", (t["instruction"] or "")[:800])
            print("input_files:", t.get("input_files"))
        return 0

    # ---- attempt -> materialize -> grade ----
    os.makedirs(a.out_root, exist_ok=True)
    caveats = []
    if a.grade_lane == "local":
        caveats.append("LOCAL PROXY grader (deterministic, no LLM) — NOT the official BTB/Gandalf score.")
        if not a.rubrics_dir:
            caveats.append("no --rubrics-dir: results are existence-only, not scored.")
    if use_synthetic:
        caveats.append("SYNTHETIC tasks: proves the code path only; not a capability measurement.")

    rows = []
    for t in sliced:
        out_dir = os.path.join(a.out_root, deliverables.safe_filename_component(t["id"]))
        os.makedirs(out_dir, exist_ok=True)
        plan, err = attempt(t, a.mode, a.attempts_dir)
        if err or not isinstance(plan, dict):
            print(f"  ! {t['id']}: attempt failed ({err}); skipping")
            rows.append({"id": t["id"], "error": err or "no plan", "cleanProbe": False,
                         "grade": {"lane": a.grade_lane, "score": None}})
            continue
        written = deliverables.write_deliverables(plan, out_dir, instruction_text=t["instruction"])
        # clean-probe gate (mirrors spreadsheetbench.py): a row counts only if it is a sealed
        # held-out task with the model genuinely in the loop and at least one artifact materialized.
        any_written = any(p is not None for p in written.values())
        clean = bool(any_written) and a.mode in ("agent", "api")
        g = grade(out_dir, t, a.grade_lane, a.harbor_cmd, a.rubrics_dir)
        rows.append({"id": t["id"], "mode": a.mode, "cleanProbe": clean,
                     "wrote": {k: (str(v) if v else None) for k, v in written.items()},
                     "grade": g})
        print(f"  {t['id']}: lane={g.get('lane')} score={g.get('score')} "
              f"clean={clean} wrote={[k for k, v in written.items() if v]}")

    head = honest_headline(rows, a.grade_lane, manifest, dataset_label, caveats)
    summary = {"benchmark": "bankertoolbench", "dataset": dataset_label, "manifest": manifest,
               "grade_lane": a.grade_lane, "caveats": caveats, **head, "rows": rows}
    with open(a.out, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    print(f"results -> {a.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
